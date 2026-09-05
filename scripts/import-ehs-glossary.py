#!/usr/bin/env python3
"""Import the user-maintained EHS glossary workbook into site JSON.

Requires openpyxl. The workbook remains the only content source; this script
normalizes blank cells and enforces the release data contract.
"""

from __future__ import annotations

from collections import Counter
from datetime import date
import json
from pathlib import Path
import re
import sys
from urllib.parse import urlparse

from openpyxl import load_workbook


EXPECTED_SHEETS = ["首页说明", "术语总库", "高频缩写", "易混术语辨析", "来源与口径"]
EXPECTED_HEADERS = ["分类", "英文术语", "缩写", "推荐中文", "简明释义", "典型使用场景", "口径来源", "参考网址", "重要度", "序号"]


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValueError(message)


def parse_category(value: str) -> tuple[str, str]:
    match = re.fullmatch(r"(\d{2})\s+(.+)", value)
    require(bool(match), f"无法识别分类：{value!r}")
    return match.group(1), match.group(2)


def is_http_url(value: str | None) -> bool:
    return value is None or urlparse(value).scheme in {"http", "https"}


def import_workbook(input_path: Path) -> dict:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    require(workbook.sheetnames == EXPECTED_SHEETS, f"Sheet不匹配：{workbook.sheetnames}")

    home = workbook["首页说明"]
    version_match = re.search(r"更新日期：(\d{4}-\d{2}-\d{2})", str(home["A2"].value or ""))
    version = version_match.group(1) if version_match else date.today().isoformat()

    term_sheet = workbook["术语总库"]
    headers = [clean(cell.value) for cell in next(term_sheet.iter_rows(min_row=1, max_row=1))]
    require(headers == EXPECTED_HEADERS, f"术语总库表头不匹配：{headers}")

    terms = []
    categories = {}
    for row in term_sheet.iter_rows(min_row=2, values_only=True):
        source = {header: clean(value) for header, value in zip(headers, row)}
        if not any(value is not None for value in source.values()):
            continue
        category_id, category_name = parse_category(source["分类"])
        categories.setdefault(category_id, category_name)
        source_url = clean(source["参考网址"])
        abbreviation = clean(source["缩写"])
        require(abbreviation not in {57, "57"}, f"序号{source['序号']}的缩写被错误解析为57")
        require(source_url not in {57, "57"}, f"序号{source['序号']}的网址被错误解析为57")
        require(is_http_url(source_url), f"序号{source['序号']}的网址协议无效：{source_url}")
        terms.append({
            "id": int(source["序号"]),
            "categoryId": category_id,
            "categoryName": category_name,
            "english": source["英文术语"],
            "abbreviation": abbreviation,
            "chinese": source["推荐中文"],
            "definition": source["简明释义"],
            "scenario": source["典型使用场景"],
            "sourceLabel": source["口径来源"],
            "sourceUrl": source_url,
            "importance": source["重要度"],
        })

    confusion_sheet = workbook["易混术语辨析"]
    confusions = []
    for index, row in enumerate(confusion_sheet.iter_rows(min_row=2, values_only=True), start=1):
        english, chinese, reminder = (clean(value) for value in row[:3])
        if not any((english, chinese, reminder)):
            continue
        confusions.append({"id": index, "english": english, "chinese": chinese, "reminder": reminder})

    sources_sheet = workbook["来源与口径"]
    sources = []
    for row in sources_sheet.iter_rows(min_row=5, values_only=True):
        label, coverage, url = (clean(value) for value in row[:3])
        if not any((label, coverage, url)):
            continue
        require(is_http_url(url), f"来源网址协议无效：{url}")
        sources.append({"label": label, "coverage": coverage, "url": url})

    importance = Counter(term["importance"] for term in terms)
    english = Counter(term["english"] for term in terms)
    chinese = Counter(term["chinese"] for term in terms)
    abbreviation_count = sum(term["abbreviation"] is not None for term in terms)
    source_url_count = sum(term["sourceUrl"] is not None for term in terms)

    require(len(terms) == 445, f"术语数应为445，实际{len(terms)}")
    require(len(categories) == 18, f"分类数应为18，实际{len(categories)}")
    require(abbreviation_count == 196, f"缩写数应为196，实际{abbreviation_count}")
    require(len(confusions) == 24, f"辨析数应为24，实际{len(confusions)}")
    require(importance == {"核心": 289, "常用": 62, "进阶": 94}, f"重要度分布错误：{dict(importance)}")
    require(source_url_count == 161, f"非空网址应为161，实际{source_url_count}")
    require(not [key for key, count in english.items() if count > 1], "英文术语存在重复")
    require(not [key for key, count in chinese.items() if count > 1], "推荐中文存在重复")
    require(sum(term["abbreviation"] == "EAP" for term in terms) == 2, "EAP必须保留2种含义")
    require(sum(term["abbreviation"] == "EPR" for term in terms) == 2, "EPR必须保留2种含义")
    require([term["id"] for term in terms] == list(range(1, 446)), "序号必须连续且保留为稳定ID")

    category_rows = [
        {
            "id": category_id,
            "name": name,
            "count": sum(term["categoryId"] == category_id for term in terms),
        }
        for category_id, name in categories.items()
    ]
    return {
        "version": version,
        "total": len(terms),
        "categoryTotal": len(category_rows),
        "abbreviationTotal": abbreviation_count,
        "confusionTotal": len(confusions),
        "sourceUrlTotal": source_url_count,
        "categories": category_rows,
        "terms": terms,
        "confusions": confusions,
        "sources": sources,
    }


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python3 scripts/import-ehs-glossary.py INPUT.xlsx OUTPUT.json")
    input_path = Path(sys.argv[1]).resolve()
    output_path = Path(sys.argv[2]).resolve()
    data = import_workbook(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "status": "PASS",
        "output": str(output_path),
        "terms": data["total"],
        "categories": data["categoryTotal"],
        "abbreviations": data["abbreviationTotal"],
        "confusions": data["confusionTotal"],
        "sourceUrls": data["sourceUrlTotal"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
